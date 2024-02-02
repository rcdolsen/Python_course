# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook: Workbook = load_workbook(WORKBOOK_PATH)

# Nome para a planilha
sheet_name = 'Shit name'

# selecionar a planilha
worksheet: Worksheet = workbook[sheet_name]  # type: ignore

for row in worksheet.iter_rows():
    for cell in row:
        print(cell.value, end='\t')
    print()
print()

for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end='\t')
    print()
print()

for row in worksheet.iter_rows(max_col=1):
    for cell in row:
        print(cell.value, end='\t')
    print()
print()

for row in worksheet.iter_rows(max_row=1):
    for cell in row:
        print(cell.value, end='\t')
    print()
print()

# para alterar os dados
# print(worksheet['B3'].value)
# worksheet['B3'].value = 18

# outro modo
for row in worksheet.iter_rows():
    for cell in row:
        if cell.value == 'Maria':
            worksheet.cell(cell.row, 2, 17)  # type: ignore


workbook.save(WORKBOOK_PATH)
